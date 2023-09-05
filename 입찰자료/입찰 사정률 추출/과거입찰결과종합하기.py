import os, re
from openpyxl import *
from openpyxl import styles
from openpyxl.styles import numbers
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import Font, Color, colors, PatternFill
import pandas as pd
import win32com.client as win32
import time

table = ["번호","공고명","발주처","공고번호","투찰마감일시","개찰일시","구분","지역제한","업종제한","예가변동폭","1순위 업체","낙찰 금액","업체 사정률","예가 사정률","업체 투찰률","기초 금액","참여업체","메모"]
area = ["전국","광주,전남,전북","강원","경기,서울,인천","대전,충남,충북","대구,경북","부산,경남,울산","제주"]


dir = r'C:\Users\user\Desktop\입찰자료\입찰통계'
os.chdir(r'C:\Users\user\Desktop\입찰자료\입찰통계')
row, column = 2, 1

write_wb = load_workbook(r'C:\Users\user\Desktop\입찰자료\입찰통계\결과종합.xlsx', data_only = True)
write_ws = write_wb.active
write_ws.title = '결과종합'


# 회색 배경색 설정
gray_fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')

# 폰트 설정
font = Font(name='Calibri', size=10, bold=True)

for i in area:
  target = write_wb.create_sheet()
  target.append(table)
  target.title = i
  target.sheet_properties.tabColor = "1DDB16"
  for r in target:
    for cell in r:
      cell.alignment = Alignment(horizontal='center', vertical='center')
      cell.fill = gray_fill
      cell.font = font


for file_name in os.listdir():
  if file_name.endswith('.xlsx') and r'KBID' in file_name:
    print(file_name)
    load_wb = load_workbook(file_name, data_only = True)
    load_ws = load_wb['KBID']
    
    cnt = 0
    for row_data in load_ws.rows:
      if cnt>99:
        break
      
      tmp = []
      for data in row_data:
        tmp.append(data.value)
  
      if(str(row_data[0].value) == "번호"): 
        continue
      ###지역별로 시트를 분류
      ### 행 단위로 시트의 데이터를 옮기기
      elif("광주" in str(row_data[7].value) or "전남" in str(row_data[7].value) or "전북" in str(row_data[7].value)):
        write_wb["광주,전남,전북"].append(tmp)
      elif("경기" in str(row_data[7].value) or "서울" in str(row_data[7].value) or "인천" in str(row_data[7].value) ):
        write_wb["경기,서울,인천"].append(tmp)
      elif("대전"  in str(row_data[7].value) or "충남" in str(row_data[7].value) or "충북" in str(row_data[7].value)):
        write_wb["대전,충남,충북"].append(tmp)
      elif("부산" in str(row_data[7].value) or "경남" in str(row_data[7].value) or "울산" in str(row_data[7].value)):
        write_wb["부산,경남,울산"].append(tmp)
      elif("대구" in str(row_data[7].value) or "경북" in str(row_data[7].value)):
        write_wb["대구,경북"].append(tmp)
      elif("제주" in str(row_data[7].value)):
        write_wb["제주"].append(tmp)
      elif("강원" in str(row_data[7].value)):
        write_wb["강원"].append(tmp)
      elif("전국" in str(row_data[7].value)):
        write_wb["전국"].append(tmp)


      write_ws = write_wb["결과종합"]
      for data in row_data:
        if(column==1) : write_ws.cell(row, column, row-1)

        elif(column==12 or column==16):
          if(data.value==None):
            write_ws.cell(row, column, str(" "))
          else:
            write_ws.cell(row, column, data.value).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

        else : 
          if(data.value==None):
            write_ws.cell(row, column, str(" "))
          else:
            write_ws.cell(row, column, data.value).number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
        column += 1
        if(column > 18):
          column = 1
          row += 1

      cnt += 1

write_wb.save(str(row)+"결과종합_renew.xlsx")
print("완료되었습니다.")